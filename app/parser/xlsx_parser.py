import openpyxl
import re


def extract_text_from_xlsx(file) -> str:
    """
    Reads text content from an XLSX file
    """
    workbook = openpyxl.load_workbook(
        file,
        data_only=True,
    )
    sheet_texts = []

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        text_data = []

        for row in worksheet.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                text_data.append(row_text.strip())

        sheet_text = " ".join(text_data).strip()
        sheet_text = re.sub(r"\s+", " ", sheet_text)
        sheet_texts.append(sheet_text)

    return " ".join(sheet_texts)
